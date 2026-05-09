"""
EIA Short-Term Energy Outlook (STEO) monthly data client.

Generic CLI tool for fetching any STEO series. The caller decides which
series to pull. No API key required — STEO is published as a public xlsx.

The xlsx contains both historical and forecast values; this pipeline does
not distinguish them. Downstream consumers cap by current month if they
want history only.

Note on sign conventions: some STEO series (e.g. t3_stchange_world) are
published draw-positive (consumption − production). This pipeline records
raw values; consumers must apply the sign convention they need.

Usage:
    python steo.py --series pasc_oecd_t3 --output /path/to/output.parquet
    python steo.py --series pasc_oecd_t3 --series papr_world --output /path/to/dir/
    python steo.py --batch batch.json
    python steo.py --series pasc_oecd_t3 --output out.parquet --sheet 3atab
"""

import argparse
import io
import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import requests
import structlog
from dotenv import load_dotenv

from _env import load_repo_env; load_repo_env()

log = structlog.get_logger()

STEO_URL = "https://www.eia.gov/outlooks/steo/xls/STEO_m.xlsx"
DEFAULT_SHEET = "3atab"

MONTH_MAP = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}


def download_workbook(url: str = STEO_URL) -> bytes:
    resp = requests.get(url, allow_redirects=True, timeout=60)
    resp.raise_for_status()
    return resp.content


def parse_dates(years_row: tuple, months_row: tuple) -> list[str | None]:
    """Walk parallel year/month header rows and emit YYYY-MM-01 strings.

    Year cells are sparse (one per January); months repeat each year.
    """
    dates: list[str | None] = []
    current_year: int | None = None
    n = min(len(years_row), len(months_row))
    for i in range(2, n):
        yr, mn = years_row[i], months_row[i]
        if yr is not None:
            try:
                current_year = int(yr)
            except (ValueError, TypeError):
                pass
        if mn is not None and current_year is not None:
            ms = str(mn)[:3]
            if ms in MONTH_MAP:
                dates.append(f"{current_year}-{MONTH_MAP[ms]:02d}-01")
                continue
        dates.append(None)
    return dates


def extract_series(content: bytes, series_ids: list[str], sheet: str) -> dict[str, pd.DataFrame]:
    """Parse the workbook once and extract requested series.

    Returns {series_id: DataFrame[date, value]}. Series IDs are case-insensitive
    in STEO conventions; we normalize to lowercase for matching.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    finally:
        wb.close()

    if len(all_rows) < 4:
        raise ValueError(f"sheet {sheet!r} too short ({len(all_rows)} rows) — STEO format may have changed")

    dates = parse_dates(all_rows[2], all_rows[3])
    wanted = {sid.lower() for sid in series_ids}
    out: dict[str, pd.DataFrame] = {sid: pd.DataFrame(columns=["date", "value"]) for sid in series_ids}
    seen: set[str] = set()

    for row in all_rows:
        if not row or not row[0]:
            continue
        sid_raw = str(row[0]).strip()
        sid_key = sid_raw.lower()
        if sid_key not in wanted or sid_key in seen:
            continue
        seen.add(sid_key)

        vals = list(row[2 : 2 + len(dates)])
        records = []
        for d, v in zip(dates, vals):
            if d is None or v is None:
                continue
            try:
                records.append((d, float(v)))
            except (ValueError, TypeError):
                continue
        df = pd.DataFrame(records, columns=["date", "value"])
        if not df.empty:
            df["date"] = pd.to_datetime(df["date"])
            df = df.sort_values("date").reset_index(drop=True)
        # Find the original-cased requested ID to key the result
        for orig in series_ids:
            if orig.lower() == sid_key:
                out[orig] = df
                break

    missing = [sid for sid in series_ids if sid.lower() not in seen]
    if missing:
        log.warning("steo_series_missing", series=missing, sheet=sheet)

    return out


def write_parquet(df: pd.DataFrame, path: Path) -> None:
    """Write DataFrame to parquet, merging with existing data on date."""
    if path.exists():
        existing = pd.read_parquet(path)
        df = pd.concat([existing, df]).drop_duplicates(subset=["date"], keep="last")
        df = df.sort_values("date").reset_index(drop=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, index=False, engine="pyarrow")


def resolve_output(output: str, series_id: str) -> Path:
    p = Path(output)
    if p.is_dir() or output.endswith("/"):
        return p / f"{series_id}.parquet"
    return p


def main():
    parser = argparse.ArgumentParser(description="Fetch EIA STEO monthly series")
    parser.add_argument("--series", action="append", help="STEO series ID (repeatable)")
    parser.add_argument("--output", help="Output parquet path (file or directory)")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Workbook sheet (default: {DEFAULT_SHEET})")
    parser.add_argument("--batch", help="JSON file with [{series, output, sheet?}, ...] for batch mode")
    args = parser.parse_args()

    work: list[tuple[str, Path, str]] = []
    if args.batch:
        with open(args.batch) as f:
            items = json.load(f)
        for item in items:
            work.append((item["series"], Path(item["output"]), item.get("sheet", DEFAULT_SHEET)))
    elif args.series:
        if not args.output:
            log.error("no_output", hint="--output is required when using --series")
            sys.exit(1)
        for sid in args.series:
            work.append((sid, resolve_output(args.output, sid), args.sheet))
    else:
        log.error("no_series", hint="Provide --series or --batch")
        sys.exit(1)

    log.info("steo_fetch", series=[s for s, _, _ in work], count=len(work))
    try:
        content = download_workbook()
    except requests.RequestException as e:
        log.error("download_failed", error=str(e))
        sys.exit(1)

    # Group by sheet so we parse each sheet at most once
    by_sheet: dict[str, list[tuple[str, Path]]] = {}
    for sid, out_path, sheet in work:
        by_sheet.setdefault(sheet, []).append((sid, out_path))

    had_failure = False
    for sheet, items in by_sheet.items():
        series_ids = [s for s, _ in items]
        try:
            results = extract_series(content, series_ids, sheet)
        except (KeyError, ValueError) as e:
            log.error("extract_failed", sheet=sheet, error=str(e))
            had_failure = True
            continue

        for sid, out_path in items:
            df = results.get(sid, pd.DataFrame())
            if df.empty:
                log.warning("no_data", series=sid, sheet=sheet)
                continue
            write_parquet(df, out_path)
            log.info("written", series=sid, sheet=sheet, output=str(out_path), rows=len(df))

    if had_failure:
        sys.exit(1)


if __name__ == "__main__":
    main()
